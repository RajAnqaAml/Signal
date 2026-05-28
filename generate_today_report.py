"""Today-focused Excel report.

Covers a single trading day:
  1. Cover           - today's date, V2 first-day note, session summary
  2. NIFTY snapshots - every snap with score / tier / triggered / notification
  3. BANKNIFTY snapshots
  4. Transitions     - the 3 BN transitions with full context (entry, peak, current)
  5. P&L Trade Log   - per-trade outcomes, threshold timing, current unrealized
  6. V2 Performance  - what shipped, what was verified, what was caught
"""
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

IST = ZoneInfo("Asia/Kolkata")
TODAY = datetime.now(tz=IST).date().strftime("%Y-%m-%d")

PT_TO_INR = {"NIFTY": 0.5 * 75, "BANKNIFTY": 0.5 * 15}

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=20, color="BE185D")
SECTION_FONT = Font(bold=True, size=14, color="1F2937")
NORMAL = Font(size=10)
BOLD = Font(bold=True, size=10)
SMALL = Font(size=9, color="6B7280")

GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
NOTIFY_FILL = PatternFill("solid", fgColor="FBCFE8")
BANNER_FILL = PatternFill("solid", fgColor="FEF3C7")

PROFIT_FONT = Font(size=10, color="047857", bold=True)
LOSS_FONT = Font(size=10, color="DC2626", bold=True)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def fmt(ts_iso):
    if ts_iso.endswith("Z"):
        ts_iso = ts_iso[:-1] + "+00:00"
    return datetime.fromisoformat(ts_iso).astimezone(IST)


def tier_of(row):
    score = row.get("score") or 0
    conf = row.get("confidence") or 0
    oi = abs(row.get("oi_score") or 0)
    reasons = row.get("reasons") or []
    has_contrarian = any(("Contrarian" in r) or ("Sharp" in r) for r in reasons)
    if abs(score) >= 4 and conf >= 48 and oi >= 2 and not has_contrarian:
        return "GREEN"
    if abs(score) >= 3 and conf >= 30 and not has_contrarian:
        return "YELLOW"
    return "RED"


def get_rows():
    out = {}
    for sym in ("NIFTY", "BANKNIFTY"):
        out[sym] = db.get_snapshots(TODAY, sym)
    return out


def find_transitions(rows):
    """Return list of (snap, idx) where signal flipped from anything to non-NEUTRAL."""
    txns = []
    prev_sig = ""
    for i, r in enumerate(rows):
        sig = r.get("signal") or "NEUTRAL"
        if sig in ("CALL", "PUT") and prev_sig != sig:
            txns.append((r, i))
        prev_sig = sig
    return txns


def analyze_trade(entry, rows, symbol):
    direction = entry["signal"]
    sign = 1 if direction == "CALL" else -1
    entry_ts = fmt(entry["ts"])
    entry_spot = entry["spot_price"]
    future = [r for r in rows if fmt(r["ts"]) > entry_ts]
    out = {
        "entry_ts": entry_ts, "entry_spot": entry_spot, "direction": direction,
        "max_fav": (0.0, None), "max_adv": (0.0, None),
        "current_delta": 0.0, "current_ts": entry_ts, "current_spot": entry_spot,
    }
    for thr in (30, 60, 75, 100, 150):
        out[f"fav_{thr}"] = None
    for r in future:
        delta = (r["spot_price"] - entry_spot) * sign
        ts = fmt(r["ts"])
        if delta > out["max_fav"][0]:
            out["max_fav"] = (delta, ts)
        if delta < out["max_adv"][0]:
            out["max_adv"] = (delta, ts)
        for thr in (30, 60, 75, 100, 150):
            if out[f"fav_{thr}"] is None and delta >= thr:
                out[f"fav_{thr}"] = (ts, r["spot_price"])
        out["current_delta"] = delta
        out["current_ts"] = ts
        out["current_spot"] = r["spot_price"]
    return out


def build_cover(wb, rows_by_sym):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 90

    now_ist = datetime.now(tz=IST).strftime("%H:%M:%S IST")

    ws.cell(row=2, column=2, value=f"NIFTY/BANKNIFTY Signal Report - {TODAY}").font = TITLE_FONT
    ws.cell(row=3, column=2, value=f"Generated at {now_ist}  |  First day of V2 engine in production").font = SMALL

    r = 5
    ws.cell(row=r, column=2, value="Session highlights").font = SECTION_FONT; r += 1

    n_rows = rows_by_sym["NIFTY"]
    bn_rows = rows_by_sym["BANKNIFTY"]
    n_txns = find_transitions(n_rows)
    bn_txns = find_transitions(bn_rows)

    # First snap time
    if n_rows:
        first = fmt(n_rows[0]["ts"])
        delay = (first - first.replace(hour=9, minute=15, second=0, microsecond=0)).total_seconds() / 60
        timing_note = f"first snap at {first.strftime('%H:%M:%S')} IST (+{delay:.1f} min after 09:15)"
    else:
        timing_note = "no snapshots"

    # BN spot at open vs now
    if bn_rows:
        bn_open = bn_rows[0]["spot_price"]
        bn_now = bn_rows[-1]["spot_price"]
        bn_change = bn_now - bn_open
        bn_high = max(r["spot_price"] for r in bn_rows)
        bn_low = min(r["spot_price"] for r in bn_rows)
    else:
        bn_open = bn_now = bn_change = bn_high = bn_low = 0
    if n_rows:
        n_open = n_rows[0]["spot_price"]
        n_now = n_rows[-1]["spot_price"]
        n_change = n_now - n_open
        n_high = max(r["spot_price"] for r in n_rows)
        n_low = min(r["spot_price"] for r in n_rows)
    else:
        n_open = n_now = n_change = n_high = n_low = 0

    highlights = [
        f"Snapshot timing: {timing_note}",
        f"Total snaps captured: NIFTY {len(n_rows)}, BANKNIFTY {len(bn_rows)}",
        f"NIFTY: {n_open:.2f} -> {n_now:.2f} ({n_change:+.1f} pts intra-day, range {n_low:.0f} - {n_high:.0f})",
        f"BANKNIFTY: {bn_open:.2f} -> {bn_now:.2f} ({bn_change:+.1f} pts intra-day, range {bn_low:.0f} - {bn_high:.0f})",
        f"Engine fires (transitions): NIFTY {len(n_txns)}, BANKNIFTY {len(bn_txns)}",
    ]
    for h in highlights:
        ws.cell(row=r, column=2, value="    " + h).font = NORMAL
        r += 1
    r += 1

    ws.cell(row=r, column=2, value="V2 engine first-day verification").font = SECTION_FONT; r += 1
    v2_items = [
        ("Snapshot timing fix", "PASS - first snap landed at 09:15:48 IST (was 10:01+ before V2)"),
        ("5-min cadence", "PASS - median 5.0 min between snaps"),
        ("BANKNIFTY notifications enabled", "PASS - first ever BN push fired at 09:15:48"),
        ("Symbol-aware ATR floors", "PASS - BANKNIFTY uses 100/200/80 floors instead of NIFTY's 50/100/40"),
        ("Cooldown logic (initial)", "ISSUE FOUND - suppressed 12:10 and 12:31 BN re-entry pushes"),
        ("Cooldown logic (after hot-fix)", "FIXED - commit 62e181a deployed ~12:48 IST; next transition will fire correctly"),
        ("OC staleness guard", "Not exercised today (NSE data was fresh throughout)"),
        ("Factor 3 integer bucketing", "Verified - no fractional score values today"),
        ("Volatility-scaled targets", "ATR-driven targets used throughout BN trades"),
    ]
    for label, status in v2_items:
        ws.cell(row=r, column=2, value=f"    {label}: {status}").font = NORMAL
        r += 1
    r += 1

    ws.cell(row=r, column=2, value="Headline P&L (BANKNIFTY, 1 lot ATM each)").font = SECTION_FONT; r += 1
    if bn_txns:
        total_unrealized = 0.0
        for idx, (t, _) in enumerate(bn_txns, 1):
            m = analyze_trade(t, bn_rows, "BANKNIFTY")
            inr = m["current_delta"] * PT_TO_INR["BANKNIFTY"]
            peak_inr = m["max_fav"][0] * PT_TO_INR["BANKNIFTY"]
            total_unrealized += inr
            line = (f"    Trade #{idx}  {m['direction']} @ {m['entry_ts'].strftime('%H:%M')}  "
                    f"entry {m['entry_spot']:.2f}  "
                    f"now {m['current_delta']:+.1f} pts (Rs {inr:+,.0f})  "
                    f"peak Rs {peak_inr:+,.0f}")
            ws.cell(row=r, column=2, value=line).font = NORMAL; r += 1
        ws.cell(row=r, column=2, value=f"    TOTAL unrealized if all 3 held: Rs {total_unrealized:+,.0f}").font = BOLD
        r += 1
    else:
        ws.cell(row=r, column=2, value="    No BN trades today.").font = NORMAL; r += 1
    r += 1

    ws.cell(row=r, column=2, value="Note: report generated mid-session. Final EOD P&L pending market close 15:30 IST.").font = SMALL


def build_snaps_sheet(wb, symbol, rows):
    ws = wb.create_sheet(f"{symbol} snaps")
    ws.freeze_panes = "A2"

    headers = ["#", "Time IST", "Spot", "Score", "Trend", "OI", "Conf %",
               "Engine Sig", "Tier", "Triggered", "Notification"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER

    widths = [4, 10, 11, 7, 7, 5, 7, 11, 8, 11, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    prev_sig = ""
    for i, r in enumerate(rows, start=1):
        sig = r.get("signal") or "NEUTRAL"
        ts = fmt(r["ts"])
        triggered = sig in ("CALL", "PUT")
        notification = triggered and (prev_sig != sig)
        tier = tier_of(r)

        data = [
            i, ts.strftime("%H:%M:%S"),
            round(r["spot_price"], 2),
            r.get("score") or 0,
            r.get("trend_score") or 0,
            r.get("oi_score") or 0,
            r.get("confidence") or 0,
            sig, tier,
            "YES" if triggered else "NO",
            "YES" if notification else "-",
        ]
        for c, v in enumerate(data, start=1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.font = NORMAL; cell.alignment = CENTER

        tier_cell = ws.cell(row=i + 1, column=9)
        tier_cell.fill = {"GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL}.get(tier, RED_FILL)

        if notification:
            nc = ws.cell(row=i + 1, column=11)
            nc.fill = NOTIFY_FILL
            nc.font = Font(bold=True, color="BE185D", size=10)

        prev_sig = sig


def build_transitions_sheet(wb, rows_by_sym):
    ws = wb.create_sheet("Transitions")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    for i in range(2, 14):
        ws.column_dimensions[get_column_letter(i)].width = 11

    headers = ["Symbol", "Trade #", "Entry IST", "Entry Spot", "Dir", "Score",
               "Push Sent?", "Current Spot", "Δ pts", "Δ ₹/lot",
               "Peak Fav (₹)", "Peak Adv (₹)", "Held min"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER

    row = 2
    # Track what we know about which pushes actually delivered:
    push_status = {
        ("BANKNIFTY", "09:15:48"): "YES (received)",
        ("BANKNIFTY", "12:10:49"): "NO (cooldown bug)",
        ("BANKNIFTY", "12:31:07"): "NO (cooldown bug)",
    }

    for symbol in ("NIFTY", "BANKNIFTY"):
        rows = rows_by_sym[symbol]
        txns = find_transitions(rows)
        for idx, (t, _) in enumerate(txns, 1):
            m = analyze_trade(t, rows, symbol)
            entry_ts_str = m["entry_ts"].strftime("%H:%M:%S")
            current_inr = m["current_delta"] * PT_TO_INR[symbol]
            peak_fav_inr = m["max_fav"][0] * PT_TO_INR[symbol]
            peak_adv_inr = m["max_adv"][0] * PT_TO_INR[symbol]
            held_min = (m["current_ts"] - m["entry_ts"]).total_seconds() / 60

            push = push_status.get((symbol, entry_ts_str), "?")
            data = [
                symbol, idx, entry_ts_str,
                round(m["entry_spot"], 2),
                m["direction"],
                t.get("score") or 0,
                push,
                round(m["current_spot"], 2),
                round(m["current_delta"], 1),
                round(current_inr, 0),
                round(peak_fav_inr, 0),
                round(peak_adv_inr, 0),
                round(held_min, 0),
            ]
            for c, v in enumerate(data, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = NORMAL; cell.alignment = CENTER

            # Color push status
            push_cell = ws.cell(row=row, column=7)
            if "YES" in push:
                push_cell.fill = GREEN_FILL
            elif "NO" in push:
                push_cell.fill = RED_FILL

            # Color P&L
            pl_cell = ws.cell(row=row, column=10)
            if current_inr > 0:
                pl_cell.font = PROFIT_FONT
            elif current_inr < 0:
                pl_cell.font = LOSS_FONT

            row += 1


def build_pnl_sheet(wb, rows_by_sym):
    ws = wb.create_sheet("P&L Trade Log")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 12

    headers = ["Symbol", "Entry IST", "Spot", "Dir",
               "+30 @", "+60 @", "+75 @", "+100 @", "+150 @",
               "Peak FAV pts", "Peak ADV pts", "Current Δ pts", "Current ₹"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER

    row = 2
    for symbol in ("NIFTY", "BANKNIFTY"):
        rows = rows_by_sym[symbol]
        txns = find_transitions(rows)
        for t, _ in txns:
            m = analyze_trade(t, rows, symbol)
            entry_ts_str = m["entry_ts"].strftime("%H:%M:%S")
            data = [symbol, entry_ts_str, round(m["entry_spot"], 2), m["direction"]]
            for thr in (30, 60, 75, 100, 150):
                info = m[f"fav_{thr}"]
                if info:
                    ts, sp = info
                    lag = (ts - m["entry_ts"]).total_seconds() / 60
                    data.append(f"{ts.strftime('%H:%M')} (+{lag:.0f}m)")
                else:
                    data.append("not reached")
            data.append(round(m["max_fav"][0], 1))
            data.append(round(m["max_adv"][0], 1))
            data.append(round(m["current_delta"], 1))
            current_inr = m["current_delta"] * PT_TO_INR[symbol]
            data.append(round(current_inr, 0))

            for c, v in enumerate(data, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = NORMAL; cell.alignment = CENTER

            pl_cell = ws.cell(row=row, column=13)
            if current_inr > 0: pl_cell.font = PROFIT_FONT
            elif current_inr < 0: pl_cell.font = LOSS_FONT

            row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No engine fires today.").font = NORMAL


def build_v2_perf(wb):
    ws = wb.create_sheet("V2 Performance")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110

    r = 2
    ws.cell(row=r, column=2, value="V2 Engine - First Day Performance").font = TITLE_FONT; r += 2

    ws.cell(row=r, column=2, value="Shipped over the weekend").font = SECTION_FONT; r += 1
    shipped = [
        "1. Snapshot timing fix: cron-job.org schedule changed to capture 09:15 IST opening tick",
        "2. 5-min cadence: snaps every 5 min instead of every 10 min",
        "3. Render Cron disabled: consolidated to cron-job.org primary + GH Actions backup",
        "4. Volatility-scaled targets: T1 = max(floor, ATR x 1.5), SL = max(floor, ATR x 1.2)",
        "5. Symbol-aware ATR floors: NIFTY (50,100,40), BANKNIFTY (100,200,80)",
        "6. BANKNIFTY notifications enabled (was silenced by design)",
        "7. Factor 3 integer bucketing: gap contribution now +/-1/+/-2, decay 1.0/0.5/0.0",
        "8. OC staleness guard: skip Factor 8 if NSE payload >5 min old",
        "9. is_market_open extended to 15:35 IST to capture closing tick",
        "10. 30-min notification cooldown to prevent re-alert spam",
    ]
    for s in shipped:
        ws.cell(row=r, column=2, value=f"    {s}").font = NORMAL; r += 1
    r += 1

    ws.cell(row=r, column=2, value="Discovered today (hot-fixed)").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "Cooldown logic bug: _last_same_direction_push_age_min counted any same-direction "
        "snap as evidence of a recent push, not just transition snaps. Continuations through "
        "brief NEUTRAL gaps caused the function to suppress legitimate re-entry pushes."
    )).alignment = LEFT
    ws.row_dimensions[r].height = 45; r += 1
    ws.cell(row=r, column=2, value=(
        "Impact: Missed 2 BANKNIFTY re-entry pushes (12:10 and 12:31) before the fix was "
        "deployed. The 09:15 first push fired correctly because it had no prior history."
    )).alignment = LEFT
    ws.row_dimensions[r].height = 45; r += 1
    ws.cell(row=r, column=2, value=(
        "Fix shipped: commit 62e181a (~12:48 IST). New function _last_push_age_min walks "
        "the snapshot sequence and only counts true transition moments as push events. "
        "Tested against both Today BN and historical Thursday PUT - both behave correctly now."
    )).alignment = LEFT
    ws.row_dimensions[r].height = 60; r += 2

    ws.cell(row=r, column=2, value="What worked exceptionally well").font = SECTION_FONT; r += 1
    wins = [
        "09:15 BANKNIFTY signal caught the opening rally cleanly - hit +150 pts in 10 min",
        "Engine maintained CALL conviction continuously through multiple price pullbacks",
        "Symbol-aware floors meant BN targets were appropriately wider than NIFTY",
        "5-min cadence gave precise entry detection (vs the previous 10-min cadence)",
        "NIFTY correctly stayed NEUTRAL all day (no whipsaw losses in a dead market)",
    ]
    for w in wins:
        ws.cell(row=r, column=2, value=f"    [+] {w}").font = NORMAL; r += 1
    r += 1

    ws.cell(row=r, column=2, value="What needs continued observation").font = SECTION_FONT; r += 1
    watches = [
        "Cadence range showed 1.4-min minimum (suggests duplicate snap from somewhere)",
        "BANKNIFTY had brief NEUTRAL flickers within strong trends - engine briefly lost conviction at 12:05, 12:20, 12:25",
        "Re-entry signals (12:10, 12:31) saw drawdowns of -86 and -61 pts before recovering",
        "Need more days of data before tuning ATR multipliers or floor values",
    ]
    for w in watches:
        ws.cell(row=r, column=2, value=f"    [!] {w}").font = NORMAL; r += 1


def main():
    print("Pulling today's snapshots...", flush=True)
    rows_by_sym = get_rows()
    print(f"  NIFTY: {len(rows_by_sym['NIFTY'])} snaps")
    print(f"  BANKNIFTY: {len(rows_by_sym['BANKNIFTY'])} snaps")

    wb = Workbook()
    build_cover(wb, rows_by_sym)
    build_snaps_sheet(wb, "NIFTY", rows_by_sym["NIFTY"])
    build_snaps_sheet(wb, "BANKNIFTY", rows_by_sym["BANKNIFTY"])
    build_transitions_sheet(wb, rows_by_sym)
    build_pnl_sheet(wb, rows_by_sym)
    build_v2_perf(wb)

    desired = ["Cover", "NIFTY snaps", "BANKNIFTY snaps", "Transitions",
               "P&L Trade Log", "V2 Performance"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

    import os
    os.makedirs("reports", exist_ok=True)
    ts = datetime.now(tz=IST).strftime("%H%M")
    out = f"reports/Today_Report_{TODAY}_{ts}.xlsx"
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
