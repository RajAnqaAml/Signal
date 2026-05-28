"""Excel report: time-to-threshold analysis for every engine-fired snap
across all available trading days (2026-05-20 through 2026-05-22).

For each fire, records the first snapshot at which the price crossed the
favorable / adverse pt-thresholds in the signal direction, plus peak
favorable and peak adverse for the rest of the day.

Sheets:
  1. Cover                          - methodology and caveats
  2. Notifications - NIFTY          - the high-signal-quality subset (NIFTY only)
  3. All Fires - NIFTY              - every NIFTY snap where engine fired
  4. Notifications - BANKNIFTY      - BN subset (push silenced live; included for comparison)
  5. All Fires - BANKNIFTY          - every BN snap where engine fired
  6. Hit-rate Summary               - what % of fires hit each threshold + median time

Threshold ladder:
  Favorable: 15, 30, 45, 60, 75 pts
  Adverse  : 15, 30, 45, 60 pts
"""
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

IST = ZoneInfo("Asia/Kolkata")

FAV_THRESHOLDS = [15, 30, 45, 60, 75]
ADV_THRESHOLDS = [15, 30, 45, 60]
PT_TO_INR = 0.50 * 75  # 37.50

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=18, color="BE185D")
SECTION_FONT = Font(bold=True, size=13, color="1F2937")
NORMAL_FONT = Font(size=10)
SMALL_FONT = Font(size=9, color="6B7280")

GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
NOTIFY_FILL = PatternFill("solid", fgColor="FBCFE8")
HIT_FILL = PatternFill("solid", fgColor="DCFCE7")  # light green for "reached"
MISS_FILL = PatternFill("solid", fgColor="F3F4F6")  # gray for "never"

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def fmt(ts_iso):
    if ts_iso.endswith("Z"):
        ts_iso = ts_iso[:-1] + "+00:00"
    return datetime.fromisoformat(ts_iso).astimezone(IST)


def tier_of(row):
    score = row.get("score") or 0
    conf = row.get("confidence") or 0
    oi = abs(row.get("oi_score") or 0)
    reasons = row.get("reasons") or []
    has_contrarian = any(("Contrarian" in r) or ("Sharp" in r) for r in reasons)
    if abs(score) >= 4 and conf >= 48 and oi >= 2 and not has_contrarian:
        return "GREEN"
    if abs(score) >= 3 and conf >= 30 and not has_contrarian:
        return "YELLOW"
    return "RED"


def compute_milestones(entry_row, future_rows, direction):
    """For one entry, walk future snaps and find first time each threshold
    is crossed in favorable / adverse direction. Returns dict."""
    entry_ts = fmt(entry_row["ts"])
    entry_spot = entry_row["spot_price"]
    sign = 1 if direction == "CALL" else -1

    out = {
        "max_fav_pts": 0.0, "max_fav_ts": None,
        "max_adv_pts": 0.0, "max_adv_ts": None,
    }
    for thr in FAV_THRESHOLDS:
        out[f"fav_{thr}_min"] = None
    for thr in ADV_THRESHOLDS:
        out[f"adv_{thr}_min"] = None

    for r in future_rows:
        delta = (r["spot_price"] - entry_spot) * sign
        ts = fmt(r["ts"])
        mins = (ts - entry_ts).total_seconds() / 60
        if delta > out["max_fav_pts"]:
            out["max_fav_pts"] = delta
            out["max_fav_ts"] = ts
        if delta < out["max_adv_pts"]:
            out["max_adv_pts"] = delta
            out["max_adv_ts"] = ts
        for thr in FAV_THRESHOLDS:
            if out[f"fav_{thr}_min"] is None and delta >= thr:
                out[f"fav_{thr}_min"] = mins
        for thr in ADV_THRESHOLDS:
            if out[f"adv_{thr}_min"] is None and delta <= -thr:
                out[f"adv_{thr}_min"] = mins
    return out


def build_cover(wb, dates_covered, notif_counts, fires_counts):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    ws.cell(row=2, column=2, value="NIFTY Signal Engine - Time-to-Threshold Report").font = TITLE_FONT
    ws.cell(row=3, column=2, value=f"Generated: {datetime.now(tz=IST).strftime('%Y-%m-%d %H:%M IST')}").font = SMALL_FONT

    r = 5
    ws.cell(row=r, column=2, value="What this report answers").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "For every engine fire (signal != NEUTRAL), how fast does the price reach +/-15, "
        "+/-30, +/-45, +/-60, +/-75 pts in the signal direction? Lets you pick a target "
        "value (and stop value) based on observed price velocity, instead of a guess."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 50
    r += 1

    ws.cell(row=r, column=2, value=f"Trading days covered: {', '.join(dates_covered)}").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value=f"NIFTY notifications: {notif_counts['NIFTY']}   |   "
                                    f"NIFTY total fires: {fires_counts['NIFTY']}").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value=f"BANKNIFTY notifications: {notif_counts['BANKNIFTY']}   |   "
                                    f"BANKNIFTY total fires: {fires_counts['BANKNIFTY']}").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Resolution caveat").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "Snapshot cadence during these 3 days was ~10 min. Therefore 'time to reach +30 pts' "
        "is rounded UP to the next snapshot. Intra-bar price touches are invisible -- the "
        "actual time to reach a threshold could be up to ~10 min faster than what this report "
        "shows. From 2026-05-25 onwards we move to 5-min cadence which will halve this error bar."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 70
    r += 1

    ws.cell(row=r, column=2, value="Sheet guide").font = SECTION_FONT; r += 1
    guide = [
        "Notifications - NIFTY     : the snaps where a phone push would have fired (TRANSITION-only)",
        "All Fires - NIFTY         : every NIFTY snap where engine returned CALL or PUT (including continuations)",
        "Notifications - BANKNIFTY : BN notifications -- push is silenced live but kept here for comparison",
        "All Fires - BANKNIFTY     : every BN fire",
        "Hit-rate Summary          : what % of fires reached each threshold + median time-to-reach",
    ]
    for line in guide:
        ws.cell(row=r, column=2, value="    " + line).font = NORMAL_FONT
        r += 1
    r += 1

    ws.cell(row=r, column=2, value="How to read each row").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "Each row is one engine fire. The +N columns show how many MINUTES from entry "
        "until the price first crossed +N pts in the signal direction. 'never' means the "
        "threshold was not reached before market close. Green cells = threshold hit. "
        "Gray cells = never reached. Peak FAV / ADV columns show the maximum the trade "
        "ever ran in each direction, with the time it peaked."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 60


def build_detail_sheet(wb, sheet_name, fires_by_day):
    """fires_by_day: dict date -> list of (snap_row, milestones, is_notification)"""
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A2"

    headers = ["Date", "#", "Time IST", "Spot", "Score", "Conf %", "Engine Sig",
               "Tier", "Notification"]
    for thr in FAV_THRESHOLDS:
        headers.append(f"+{thr} pts (min)")
    for thr in ADV_THRESHOLDS:
        headers.append(f"-{thr} pts (min)")
    headers.extend(["Peak FAV pts", "Peak FAV time", "Peak FAV at +min",
                    "Peak ADV pts", "Peak ADV time", "Peak ADV at +min"])

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER

    widths = [12, 4, 9, 10, 7, 7, 11, 8, 12]
    widths += [11] * len(FAV_THRESHOLDS)
    widths += [11] * len(ADV_THRESHOLDS)
    widths += [11, 12, 13, 11, 12, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for date_str in sorted(fires_by_day.keys()):
        for idx, (snap, ms, is_notif) in enumerate(fires_by_day[date_str], start=1):
            entry_ts = fmt(snap["ts"])
            tier = tier_of(snap)
            data = [
                date_str, idx, entry_ts.strftime("%H:%M:%S"),
                round(snap["spot_price"], 2),
                snap.get("score") or 0,
                snap.get("confidence") or 0,
                snap.get("signal") or "NEUTRAL",
                tier,
                "YES" if is_notif else "-",
            ]
            for thr in FAV_THRESHOLDS:
                v = ms[f"fav_{thr}_min"]
                data.append(round(v, 0) if v is not None else "never")
            for thr in ADV_THRESHOLDS:
                v = ms[f"adv_{thr}_min"]
                data.append(round(v, 0) if v is not None else "never")
            mf_pts = ms["max_fav_pts"]
            mf_ts = ms["max_fav_ts"]
            ma_pts = ms["max_adv_pts"]
            ma_ts = ms["max_adv_ts"]
            data.extend([
                round(mf_pts, 1) if mf_ts else 0,
                mf_ts.strftime("%H:%M") if mf_ts else "-",
                round((mf_ts - entry_ts).total_seconds() / 60, 0) if mf_ts else "-",
                round(ma_pts, 1) if ma_ts else 0,
                ma_ts.strftime("%H:%M") if ma_ts else "-",
                round((ma_ts - entry_ts).total_seconds() / 60, 0) if ma_ts else "-",
            ])

            for c, v in enumerate(data, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = NORMAL_FONT
                cell.alignment = CENTER

            # Tier coloring (col 8)
            tier_cell = ws.cell(row=row, column=8)
            tier_cell.fill = {"GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL}.get(tier, RED_FILL)

            # Notification highlight (col 9)
            if is_notif:
                nc = ws.cell(row=row, column=9)
                nc.fill = NOTIFY_FILL
                nc.font = Font(bold=True, color="BE185D", size=10)

            # Color the threshold cells
            col_offset = 9
            for i, thr in enumerate(FAV_THRESHOLDS, start=1):
                cc = ws.cell(row=row, column=col_offset + i)
                if cc.value == "never":
                    cc.fill = MISS_FILL
                else:
                    cc.fill = HIT_FILL
            col_offset = 9 + len(FAV_THRESHOLDS)
            for i, thr in enumerate(ADV_THRESHOLDS, start=1):
                cc = ws.cell(row=row, column=col_offset + i)
                if cc.value == "never":
                    cc.fill = MISS_FILL
                else:
                    cc.fill = RED_FILL  # adverse hit is bad

            row += 1


def build_summary(wb, fires_by_symbol):
    """fires_by_symbol: {symbol: {date: [(snap, ms, is_notif)]}}"""
    ws = wb.create_sheet("Hit-rate Summary")
    ws.freeze_panes = "A2"
    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["A"].width = 18

    headers = ["Scope", "Symbol", "Fires (n)",
               "Hit +15 %", "Hit +30 %", "Hit +45 %", "Hit +60 %", "Hit +75 %",
               "Med min +30", "Med min +45", "Hit -30 %"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER

    r = 2
    for scope_label, scope_filter in (("All fires", lambda is_n: True),
                                       ("Notifications only", lambda is_n: is_n)):
        for symbol in ("NIFTY", "BANKNIFTY"):
            by_day = fires_by_symbol[symbol]
            flat = [(s, m, n) for date in by_day for (s, m, n) in by_day[date] if scope_filter(n)]
            n = len(flat)
            if n == 0:
                continue
            def pct(thr_key):
                hits = sum(1 for _, m, _ in flat if m[thr_key] is not None)
                return f"{100 * hits / n:.0f}%"
            def median(thr_key):
                vals = sorted(m[thr_key] for _, m, _ in flat if m[thr_key] is not None)
                if not vals:
                    return "-"
                return f"{vals[len(vals)//2]:.0f}"
            data = [scope_label, symbol, n,
                    pct("fav_15_min"), pct("fav_30_min"), pct("fav_45_min"),
                    pct("fav_60_min"), pct("fav_75_min"),
                    median("fav_30_min"), median("fav_45_min"),
                    pct("adv_30_min")]
            for c, v in enumerate(data, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = NORMAL_FONT
                cell.alignment = CENTER
            r += 1


def collect_fires(symbol, dates):
    """For a symbol, return:
       fires_by_day: {date: [(snap_row, milestones, is_notification), ...]}
       notif_count, fires_count
    """
    fires_by_day = {}
    cli = db.client(service=True)
    notif_count = 0
    fires_count = 0
    for d in dates:
        rows = db.get_snapshots(d, symbol)
        if not rows:
            continue
        prev_sig = ""
        day_fires = []
        for i, r in enumerate(rows):
            sig = r.get("signal") or "NEUTRAL"
            triggered = sig in ("CALL", "PUT")
            is_notif = triggered and (prev_sig != sig)
            if triggered:
                fires_count += 1
                future = rows[i+1:]
                direction = sig
                ms = compute_milestones(r, future, direction)
                day_fires.append((r, ms, is_notif))
                if is_notif:
                    notif_count += 1
            prev_sig = sig
        if day_fires:
            fires_by_day[d] = day_fires
    return fires_by_day, notif_count, fires_count


def main():
    dates = ["2026-05-20", "2026-05-21", "2026-05-22"]
    print("Pulling snapshots and computing milestones...", flush=True)

    fires_by_symbol = {}
    notif_counts = {}
    fires_counts = {}
    for symbol in ("NIFTY", "BANKNIFTY"):
        f, n, fc = collect_fires(symbol, dates)
        fires_by_symbol[symbol] = f
        notif_counts[symbol] = n
        fires_counts[symbol] = fc
        print(f"  {symbol}: {fc} fires, {n} notifications across {len(f)} days")

    wb = Workbook()
    build_cover(wb, dates, notif_counts, fires_counts)

    # Notification subset for each symbol
    for sym in ("NIFTY", "BANKNIFTY"):
        notif_only = {d: [(s, m, n) for s, m, n in fires_by_symbol[sym][d] if n]
                      for d in fires_by_symbol[sym] if any(n for _, _, n in fires_by_symbol[sym][d])}
        build_detail_sheet(wb, f"Notifications - {sym}", notif_only)
        build_detail_sheet(wb, f"All Fires - {sym}", fires_by_symbol[sym])

    build_summary(wb, fires_by_symbol)

    desired = ["Cover",
               "Notifications - NIFTY", "All Fires - NIFTY",
               "Notifications - BANKNIFTY", "All Fires - BANKNIFTY",
               "Hit-rate Summary"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

    import os
    os.makedirs("reports", exist_ok=True)
    out_path = f"reports/Time_to_Threshold_Report_{datetime.now(tz=IST).strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
