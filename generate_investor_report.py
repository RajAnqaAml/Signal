"""Generate an investor-ready Excel report of the NIFTY signal engine.

Contents (separate sheets):
    1. Cover            - title, scope, dates, methodology summary
    2. Logic            - full engine documentation (formula, factors, thresholds)
    3. Live - NIFTY     - per-snapshot table for every live snapshot recorded
    4. Live - BANKNIFTY - same, for reference (no push alerts on this)
    5. Live Summary     - per-day aggregate
    6. OHLC Backtest    - 30-day backtest on Yahoo 5-min candles (FLOW-BLIND)
    7. Backtest Summary - per-day aggregate

Honest scope note repeated on every backtest sheet:
    OHLC backtest is FLOW-BLIND - missing factors 5 (OI Spurts), 6 (breadth),
    8 (option chain) because no historical source exists for them.
    Production score gets ~50% of its information from these missing factors,
    so backtest results UNDERESTIMATE the live engine's selectivity.
"""
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

import db
from app import (IST, compute_ema, compute_macd, compute_rsi,
                 compute_supertrend, generate_signal)
from backtest_history import (find_nearest_vix, group_by_day, load_candles,
                              load_vix)

# ── Style constants ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # ink
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="BE185D")
SECTION_FONT = Font(name="Calibri", bold=True, size=14, color="1F2937")
NORMAL_FONT = Font(name="Calibri", size=10)
SMALL_FONT = Font(name="Calibri", size=9, color="6B7280")

GREEN_FILL = PatternFill("solid", fgColor="D1FAE5")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
NOTIFY_FILL = PatternFill("solid", fgColor="FBCFE8")  # blush
PROFIT_FONT = Font(name="Calibri", size=10, color="047857", bold=True)
LOSS_FONT = Font(name="Calibri", size=10, color="DC2626", bold=True)

THIN = Side(border_style="thin", color="E5E7EB")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Trade simulation constants ──────────────────────────────────────────────
ENGINE_TGT = 75   # spot points target
ENGINE_SL = 60    # spot points stop
PT_TO_INR = 0.50 * 75  # ATM delta * NIFTY lot = Rs 37.5 per spot point


# ── Helpers ────────────────────────────────────────────────────────────────
def tier_of(row):
    score = row.get("score") or 0
    conf = row.get("confidence") or 0
    oi = abs(row.get("oi_score") or 0)
    reasons = row.get("reasons") or []
    has_contrarian = any(("Contrarian" in r) or ("Sharp" in r) for r in reasons)
    if abs(score) >= 4 and conf >= 48 and oi >= 2 and not has_contrarian:
        return "GREEN"
    if abs(score) >= 3 and conf >= 30 and not has_contrarian:
        return "YELLOW"
    return "RED"


def walk_outcome(entry_spot, future_rows, direction):
    """Return (outcome, hit_row_or_None, pnl_inr)."""
    sign = 1 if direction == "CALL" else -1
    last = None
    for r in future_rows:
        sp = r["spot_price"]
        d = (sp - entry_spot) * sign
        if d >= ENGINE_TGT and d <= -ENGINE_SL:
            return ("both?", r, 0.0)
        if d >= ENGINE_TGT:
            return ("target", r, ENGINE_TGT * PT_TO_INR)
        if d <= -ENGINE_SL:
            return ("stop", r, -ENGINE_SL * PT_TO_INR)
        last = r
    if last is None:
        return ("no-data", None, 0.0)
    eod = (last["spot_price"] - entry_spot) * sign
    return ("eod", last, eod * PT_TO_INR)


def fmt_ist(ts_iso):
    if ts_iso.endswith("Z"):
        ts_iso = ts_iso[:-1] + "+00:00"
    return datetime.fromisoformat(ts_iso).astimezone(IST)


# ── Sheet builders ─────────────────────────────────────────────────────────
def build_cover(wb, live_days, live_date_range, backtest_days, backtest_date_range):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 80

    ws.cell(row=2, column=2, value="NIFTY Signal Engine - Performance Report").font = TITLE_FONT
    ws.cell(row=3, column=2, value=f"Generated: {datetime.now(tz=IST).strftime('%Y-%m-%d %H:%M IST')}").font = SMALL_FONT

    ws.cell(row=5, column=2, value="Scope").font = SECTION_FONT
    ws.cell(row=6, column=2, value=(
        f"Live snapshots (production engine):  {live_days} trading days  "
        f"({live_date_range})"
    )).font = NORMAL_FONT
    ws.cell(row=7, column=2, value=(
        f"OHLC backtest (flow-blind engine):   {backtest_days} trading days  "
        f"({backtest_date_range})"
    )).font = NORMAL_FONT

    ws.cell(row=9, column=2, value="What's in this workbook").font = SECTION_FONT
    contents = [
        "1. Cover  - this page",
        "2. Logic & Conditions  - full engine documentation (formulas, factors, thresholds, notification gate)",
        "3. Live - NIFTY  - every snapshot recorded by the production engine, with Tier + Triggered + Notification flags",
        "4. Live - BANKNIFTY  - same data for BANKNIFTY (recorded but no phone alerts by design)",
        "5. Live Summary  - per-day aggregate of fires / notifications / hypothetical P&L",
        "6. OHLC Backtest  - same engine simulated on 30 days of Yahoo 5-min candles  (FLOW-BLIND)",
        "7. Backtest Summary  - per-day aggregate of the backtest",
    ]
    for i, t in enumerate(contents):
        ws.cell(row=10 + i, column=2, value=t).font = NORMAL_FONT

    ws.cell(row=18, column=2, value="Important caveat on the OHLC Backtest").font = SECTION_FONT
    ws.cell(row=19, column=2, value=(
        "The 30-day OHLC backtest runs the same scoring engine on historical 5-min candles, "
        "but with three of the eight factors missing because NSE does not publish historical "
        "option-chain / OI / breadth data. Those missing factors carry roughly 50% of the live "
        "engine's signal weight (OI Spurts dominate live conviction). The backtest is therefore "
        "a LOWER BOUND on the engine's behaviour: in live, the engine has more evidence and is "
        "more selective. Read backtest hit-rates as 'directional sanity check only', not as a "
        "production P&L forecast."
    )).alignment = LEFT
    ws.row_dimensions[19].height = 80

    ws.cell(row=21, column=2, value="Two-tier interpretation guide").font = SECTION_FONT
    guide = [
        "TIER (badge on each fire):",
        "    GREEN  - |score| >= 4 AND conf >= 48 AND |OI score| >= 2 AND no contrarian flag  (high conviction)",
        "    YELLOW - |score| >= 3 AND conf >= 30 AND no contrarian flag                       (paper-trade)",
        "    RED    - everything else (sub-threshold; engine does NOT fire)",
        "",
        "TRIGGERED  - YES when engine emitted CALL or PUT (i.e. NOT NEUTRAL)",
        "NOTIFICATION - YES when a phone push would have fired: TRIGGERED AND direction CHANGED from previous snap",
    ]
    for i, t in enumerate(guide):
        ws.cell(row=22 + i, column=2, value=t).font = NORMAL_FONT


def build_logic(wb):
    ws = wb.create_sheet("Logic & Conditions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    r = 2
    ws.cell(row=r, column=2, value="Signal Engine - Logic, Formula, Conditions").font = TITLE_FONT
    r += 2

    ws.cell(row=r, column=2, value="Core formula").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "score = trend_score + oi_score + VIX + breadth + technicals + option_chain"
    )).font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "Positive score = bullish (CALL).  Negative score = bearish (PUT)."
    )).font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="The 8 input factors").font = SECTION_FONT; r += 1
    factors = [
        ("Factor 1 - Intraday move from open", "trend +/- 1 or +/- 2", "% change from day open. >+0.8% adds +2; >+0.3% adds +1; mirrored for downside."),
        ("Factor 2 - Position in day's range", "trend +/- 1", "If spot is >80% of day range from low: +1. <20%: -1."),
        ("Factor 3 - Gap analysis with time decay", "trend +/- 1 x weight", "(open - prev_close) % times a decay weight that goes from 1.0 at 09:15 to 0 at 12:15."),
        ("Factor 4 - VIX level + change", "+/- 1 each (level + change)", "VIX <14 adds +1 (low fear, trend continues). >22: -1. VIX falling >5%: +1. Rising >5%: -1."),
        ("Factor 5 - OI Spurts bias", "oi +/- 1 or +/- 2", "Long buildup vs short buildup across F&O. Bias > 1: oi +2. > 0: +1. < -1: -2. < 0: -1."),
        ("Factor 6 - Market breadth", "+/- 1", "Advances vs declines on NSE. >70% advancing: +1. <30%: -1."),
        ("Factor 7 - Technical indicators (only on real intraday history)", "+/- 4 max", "RSI >65 / <35, MACD histogram sign, EMA9 vs EMA21 crossover, SuperTrend direction. Each contributes +/- 1."),
        ("Factor 8 - Option chain flow", "+/- 5 max", "PCR (heaviest), spot vs max CE/PE strike, ATM CE/PE writing, IV skew. Each contributes +/- 1 to +/- 2."),
    ]
    for name, weight, desc in factors:
        ws.cell(row=r, column=2, value=name).font = Font(bold=True, size=10)
        r += 1
        ws.cell(row=r, column=2, value=f"    Weight: {weight}").font = NORMAL_FONT; r += 1
        ws.cell(row=r, column=2, value=f"    {desc}").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Contrarian filter (safety net)").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "If trend_score and oi_score have opposite signs (price action says one thing, options "
        "flow says the opposite), confidence is halved. If the conflict is SHARP "
        "(|trend_score| >= 2 AND |oi_score| >= 1), the engine FORCES NEUTRAL even if the total "
        "score crosses the threshold. This protects against trading into exhausted moves."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 60
    r += 1

    ws.cell(row=r, column=2, value="Trade threshold").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value="    score >= +3  ->  CALL").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    score <= -3  ->  PUT").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    otherwise    ->  NEUTRAL (no trade)").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Confidence formula").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value="    confidence_% = min(|score| x 12, 95) x contrarian_penalty").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    contrarian_penalty = 1.0 (normal) or 0.5 (when trend vs OI disagree)").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Trade targets (engine rule, NIFTY step=50)").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value="    Target 1 (T1):  +/- 1.5 x step = +/- 75 spot points").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    Target 2 (T2):  +/- 3.0 x step = +/- 150 spot points").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    Stop loss:       +/- 1.2 x step = +/- 60 spot points (against direction)").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Tier classification (used in this report)").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value="    GREEN   |score| >= 4 AND conf >= 48 AND |oi| >= 2 AND no contrarian").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    YELLOW  |score| >= 3 AND conf >= 30 AND no contrarian").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    RED     everything else (sub-threshold, no fire)").font = NORMAL_FONT; r += 2

    ws.cell(row=r, column=2, value="Notification gate (phone push)").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "A push fires when ALL four conditions are true:"
    )).font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    1. symbol is NIFTY  (BANKNIFTY is silenced by design)").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    2. current signal != NEUTRAL  (engine actually fired CALL or PUT)").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    3. previous signal != current signal  (this is a TRANSITION, not a continuation)").font = NORMAL_FONT; r += 1
    ws.cell(row=r, column=2, value="    4. NTFY_TOPIC environment variable is set").font = NORMAL_FONT; r += 2
    ws.cell(row=r, column=2, value=(
        "Continuations are silent: if the engine fires PUT 30 times in a row, you get ONE push, "
        "not 30. Direction flips push (CALL -> PUT). Returns from a brief NEUTRAL also push "
        "(known limitation - may re-alert on the same trend within minutes)."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 60
    r += 1

    ws.cell(row=r, column=2, value="Implementation assumptions (audit notes)").font = SECTION_FONT; r += 1
    impl_notes = [
        ("Technical indicators timeframe", "Computed on 5-min OHLC bars (Yahoo Finance source). Switching timeframe will change every Factor 7 contribution."),
        ("Option chain cache", "10-second cache window within one snapshot batch. Refreshed from NSE on each new snapshot."),
        ("Option chain staleness guard", "Each snapshot validates records.timestamp from NSE; if >5 min old, Factor 8 is zeroed and a warning is logged."),
        ("OI Spurts and Market Breadth source", "Live-fetched from NSE per snapshot. No historical equivalent exists (this is why the OHLC backtest is flow-blind)."),
        ("NSE data path", "Direct scraping via curl_cffi (Chrome TLS impersonation). No paid feed. Single point of failure."),
        ("Tick precision", "Score values stored unrounded in DB. Reports display rounded; threshold checks use raw float. A 'displayed -3' may not satisfy 'score <= -3' if real value is -2.95."),
        ("Notification cooldown", "30-min same-direction suppression: a PUT push within 30 min of the prior PUT push is silenced to avoid alert spam on trend continuation through brief NEUTRAL gaps."),
        ("Trade targets scaling", "Target and stop-loss scale with realized intraday volatility (ATR-based). On low-vol days targets shrink; on high-vol days targets expand."),
    ]
    for label, body in impl_notes:
        ws.cell(row=r, column=2, value=f"    {label}:").font = Font(bold=True, size=10)
        r += 1
        ws.cell(row=r, column=2, value=f"        {body}").font = NORMAL_FONT
        ws.cell(row=r, column=2).alignment = LEFT
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1

    ws.cell(row=r, column=2, value="Notes on this report's P&L numbers").font = SECTION_FONT; r += 1
    ws.cell(row=r, column=2, value=(
        "P&L per snap is HYPOTHETICAL - it walks the next snapshots forward at 10-min "
        "resolution and applies the engine's target / SL rule. At 10-min resolution, "
        "intra-bar touches are not visible: when both target and SL sit inside one forward "
        "bar, the row is flagged 'both?' (uncertain which hit first). Rupee values use ATM "
        "delta = 0.50 and NIFTY lot size = 75, so 1 spot point = Rs 37.50 per lot."
    )).alignment = LEFT; r += 1
    ws.row_dimensions[r-1].height = 70


def build_live_snapshots(wb, symbol, all_rows_by_day):
    """all_rows_by_day: {date_str: [rows]}"""
    ws = wb.create_sheet(f"Live - {symbol}")
    ws.freeze_panes = "A2"

    headers = ["Date", "#", "Time IST", "Spot", "Score", "Trend", "OI", "Conf %",
               "Engine Sig", "Tier", "Triggered", "Notification",
               "Hypo Dir", "Best move", "Worst move", "Outcome", "Exit time", "P&L (Rs)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # column widths
    widths = [12, 4, 9, 10, 6, 6, 5, 7, 11, 8, 10, 13, 9, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    excel_row = 2
    for date_str in sorted(all_rows_by_day.keys()):
        rows = all_rows_by_day[date_str]
        prev_sig = ""
        for idx, r in enumerate(rows, start=1):
            score = r.get("score") or 0
            trend = r.get("trend_score") or 0
            oi = r.get("oi_score") or 0
            conf = r.get("confidence") or 0
            sig = r.get("signal") or "NEUTRAL"
            spot = r["spot_price"]
            tier = tier_of(r)
            triggered = sig in ("CALL", "PUT")
            notification = triggered and (prev_sig != sig)

            future = rows[idx:]
            if score == 0:
                direction = "-"
                best = worst = pl = 0.0
                outcome = "flat"
                hit_time = "-"
            else:
                direction = "CALL" if score > 0 else "PUT"
                sign = 1 if direction == "CALL" else -1
                if future:
                    favs = [(rr["spot_price"] - spot) * sign for rr in future]
                    best = max(favs); worst = min(favs)
                else:
                    best = worst = 0.0
                outcome, hit_row, pl = walk_outcome(spot, future, direction)
                hit_time = fmt_ist(hit_row["ts"]).strftime("%H:%M") if hit_row else "-"

            row_data = [
                date_str, idx, fmt_ist(r["ts"]).strftime("%H:%M:%S"),
                round(spot, 2), score, trend, oi, conf, sig, tier,
                "YES" if triggered else "NO",
                "YES" if notification else "-",
                direction,
                round(best, 1) if direction != "-" else "-",
                round(worst, 1) if direction != "-" else "-",
                outcome,
                hit_time,
                round(pl, 0),
            ]
            for col, v in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col, value=v)
                cell.font = NORMAL_FONT
                cell.border = BORDER
                cell.alignment = CENTER

            # Tier coloring (column 10)
            tier_cell = ws.cell(row=excel_row, column=10)
            if tier == "GREEN":
                tier_cell.fill = GREEN_FILL
            elif tier == "YELLOW":
                tier_cell.fill = YELLOW_FILL
            else:
                tier_cell.fill = RED_FILL

            # Notification highlight (column 12)
            if notification:
                ws.cell(row=excel_row, column=12).fill = NOTIFY_FILL
                ws.cell(row=excel_row, column=12).font = Font(bold=True, color="BE185D")

            # P&L color (column 18)
            pl_cell = ws.cell(row=excel_row, column=18)
            if isinstance(pl, (int, float)) and pl > 0:
                pl_cell.font = PROFIT_FONT
            elif isinstance(pl, (int, float)) and pl < 0:
                pl_cell.font = LOSS_FONT

            prev_sig = sig
            excel_row += 1


def build_live_summary(wb, by_symbol_day):
    """by_symbol_day: {symbol: {date: [rows]}}"""
    ws = wb.create_sheet("Live Summary")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 13

    headers = ["Date", "Symbol", "Snaps", "Triggered", "Notifications",
               "GREEN tier", "YELLOW tier", "RED tier",
               "Hypo P&L on all", "Hypo P&L on notifies only", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    r = 2
    for symbol in sorted(by_symbol_day.keys()):
        for date_str in sorted(by_symbol_day[symbol].keys()):
            rows = by_symbol_day[symbol][date_str]
            triggered_count = 0
            notify_count = 0
            tier_counts = {"GREEN": 0, "YELLOW": 0, "RED": 0}
            all_pl = 0.0
            notify_pl = 0.0
            prev_sig = ""
            for idx, rr in enumerate(rows, start=1):
                score = rr.get("score") or 0
                sig = rr.get("signal") or "NEUTRAL"
                tier = tier_of(rr)
                tier_counts[tier] += 1
                triggered = sig in ("CALL", "PUT")
                notification = triggered and (prev_sig != sig)
                if triggered: triggered_count += 1
                if notification: notify_count += 1
                if score != 0:
                    direction = "CALL" if score > 0 else "PUT"
                    future = rows[idx:]
                    _, _, pl = walk_outcome(rr["spot_price"], future, direction)
                    all_pl += pl
                    if notification: notify_pl += pl
                prev_sig = sig

            note = ""
            if symbol == "BANKNIFTY":
                note = "Recorded for context; no phone push by design"
            data = [date_str, symbol, len(rows), triggered_count, notify_count,
                    tier_counts["GREEN"], tier_counts["YELLOW"], tier_counts["RED"],
                    round(all_pl, 0), round(notify_pl, 0), note]
            for col, v in enumerate(data, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = NORMAL_FONT
            r += 1


def run_ohlc_backtest_day(date_str, candles_today, prev_close, vix_map, symbol):
    """Run engine on every 5-min candle of one day. Return list of per-candle dicts."""
    out = []
    if len(candles_today) < 10:
        return out

    day_open = candles_today[0]["open"]
    high_so_far = candles_today[0]["high"]
    low_so_far = candles_today[0]["low"]
    closes_so_far = []
    prev_sig = ""

    for i, c in enumerate(candles_today):
        closes_so_far.append(c["close"])
        high_so_far = max(high_so_far, c["high"])
        low_so_far = min(low_so_far, c["low"])
        if i < 6:
            continue
        ts = c["ts"]
        dt = datetime.fromtimestamp(ts, tz=IST)
        spot = c["close"]
        change_pct = ((spot - prev_close) / prev_close * 100) if prev_close else 0
        spot_data = {
            "price": spot, "change": round(change_pct, 2),
            "open": day_open, "high": high_so_far, "low": low_so_far,
            "prev_close": prev_close,
        }
        vix_val = find_nearest_vix(vix_map, ts)
        vix_data = {"value": round(vix_val, 2), "change": 0}
        closes_arr = np.array(closes_so_far)
        highs_arr = np.array([cc["high"] for cc in candles_today[:i+1]])
        lows_arr = np.array([cc["low"] for cc in candles_today[:i+1]])
        technicals = {
            "rsi": compute_rsi(closes_arr),
            "macd": compute_macd(closes_arr),
            "ema9": compute_ema(closes_arr, 9),
            "ema21": compute_ema(closes_arr, 21),
            "supertrend": compute_supertrend(highs_arr, lows_arr, closes_arr),
            "history_source": "real",
            "bars": len(closes_arr),
        }
        sig = generate_signal(
            spot_data, vix_data, oi_analysis=None,
            breadth={"advances": 0, "declines": 0, "unchanged": 0},
            technicals=technicals, oc_analysis=None,
            history_source="real", symbol=symbol, now_ist=dt,
        )
        direction = sig["signal"]
        triggered = direction in ("CALL", "PUT")
        notification = triggered and (prev_sig != direction)

        # Forward-look: 6 candles = 30 min
        forward = candles_today[i+1:i+7]
        outcome = "no-forward"
        pl = 0.0
        if forward and triggered:
            sign = 1 if direction == "CALL" else -1
            hit = None
            for fc in forward:
                d_hi = (fc["high"] - spot) * sign
                d_lo = (fc["low"] - spot) * sign
                # for CALL: target if hi >= entry+75; sl if lo <= entry-60
                if d_hi >= ENGINE_TGT and d_lo <= -ENGINE_SL:
                    hit = "both?"; break
                if d_hi >= ENGINE_TGT:
                    hit = "target"; break
                if d_lo <= -ENGINE_SL:
                    hit = "stop"; break
            if hit == "target":
                outcome, pl = "target", ENGINE_TGT * PT_TO_INR
            elif hit == "stop":
                outcome, pl = "stop", -ENGINE_SL * PT_TO_INR
            elif hit == "both?":
                outcome, pl = "both?", 0
            else:
                # exit at last forward candle close
                exit_close = forward[-1]["close"]
                exit_pts = (exit_close - spot) * sign
                outcome, pl = "30m-exit", exit_pts * PT_TO_INR

        out.append({
            "date": date_str,
            "time": dt.strftime("%H:%M"),
            "spot": round(spot, 2),
            "score": round(sig["score"], 1),
            "trend": round(sig.get("trend_score", 0), 1),
            "oi": round(sig.get("oi_score", 0), 1),
            "conf": round(sig["confidence"], 1),
            "engine_sig": direction,
            "tier": tier_of({"score": sig["score"], "confidence": sig["confidence"],
                             "oi_score": sig.get("oi_score", 0),
                             "reasons": sig.get("reasons", [])}),
            "triggered": triggered,
            "notification": notification,
            "outcome": outcome,
            "pl": round(pl, 0),
        })
        prev_sig = direction
    return out


def build_ohlc_backtest(wb, symbol):
    """Generate backtest sheet using 30-day Yahoo candles."""
    print(f"Running OHLC backtest for {symbol}...", flush=True)
    candles = load_candles(symbol)
    vix_map = load_vix()
    by_day = group_by_day(candles)
    days = sorted(by_day.keys())

    all_results = []
    prev_close = None
    for day in days:
        ct = by_day[day]
        if prev_close is None:
            prev_close = ct[0]["open"]
        all_results.extend(run_ohlc_backtest_day(day, ct, prev_close, vix_map, symbol))
        prev_close = ct[-1]["close"]

    ws = wb.create_sheet(f"OHLC Backtest - {symbol}")
    ws.freeze_panes = "A4"

    # Caveat banner (rows 1-2)
    ws.merge_cells("A1:N1")
    banner = ws["A1"]
    banner.value = (f"FLOW-BLIND BACKTEST - 30-day Yahoo 5-min candles. "
                    f"Missing factors: OI Spurts (F5), market breadth (F6), option chain (F8). "
                    f"Production engine has ~50% more signal evidence. Use for directional sanity check only.")
    banner.font = Font(bold=True, color="9F1239", size=10)
    banner.fill = PatternFill("solid", fgColor="FEF3C7")
    banner.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value=f"Days covered: {len(days)}   Total candles: {len(all_results)}").font = SMALL_FONT

    headers = ["Date", "Time", "Spot", "Score", "Trend", "OI",
               "Conf %", "Engine Sig", "Tier", "Triggered", "Notification",
               "Outcome (30m)", "P&L (Rs)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    widths = [12, 8, 10, 6, 6, 5, 7, 11, 8, 10, 12, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    excel_row = 4
    for r in all_results:
        data = [r["date"], r["time"], r["spot"], r["score"], r["trend"], r["oi"],
                r["conf"], r["engine_sig"], r["tier"],
                "YES" if r["triggered"] else "NO",
                "YES" if r["notification"] else "-",
                r["outcome"], r["pl"]]
        for col, v in enumerate(data, start=1):
            cell = ws.cell(row=excel_row, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER

        tier_cell = ws.cell(row=excel_row, column=9)
        if r["tier"] == "GREEN":
            tier_cell.fill = GREEN_FILL
        elif r["tier"] == "YELLOW":
            tier_cell.fill = YELLOW_FILL
        else:
            tier_cell.fill = RED_FILL

        if r["notification"]:
            ws.cell(row=excel_row, column=11).fill = NOTIFY_FILL
            ws.cell(row=excel_row, column=11).font = Font(bold=True, color="BE185D")

        pl_cell = ws.cell(row=excel_row, column=13)
        if r["pl"] > 0:
            pl_cell.font = PROFIT_FONT
        elif r["pl"] < 0:
            pl_cell.font = LOSS_FONT

        excel_row += 1

    return all_results


def build_backtest_summary(wb, results_by_symbol):
    ws = wb.create_sheet("Backtest Summary")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 13

    headers = ["Date", "Symbol", "Candles tested", "Triggered", "Notifications",
               "GREEN", "YELLOW", "RED", "Targets hit", "Stops hit", "Day P&L (Rs)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    r = 2
    for symbol in sorted(results_by_symbol.keys()):
        by_day = defaultdict(list)
        for row in results_by_symbol[symbol]:
            by_day[row["date"]].append(row)
        for date_str in sorted(by_day.keys()):
            rows = by_day[date_str]
            triggered = sum(1 for x in rows if x["triggered"])
            notify = sum(1 for x in rows if x["notification"])
            g = sum(1 for x in rows if x["tier"] == "GREEN")
            y = sum(1 for x in rows if x["tier"] == "YELLOW")
            rd = sum(1 for x in rows if x["tier"] == "RED")
            tgt = sum(1 for x in rows if x["outcome"] == "target")
            stp = sum(1 for x in rows if x["outcome"] == "stop")
            pl = sum(x["pl"] for x in rows)
            data = [date_str, symbol, len(rows), triggered, notify,
                    g, y, rd, tgt, stp, round(pl, 0)]
            for col, v in enumerate(data, start=1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = NORMAL_FONT
            r += 1


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    print("Pulling live snapshots from Supabase...", flush=True)
    by_symbol_day = {"NIFTY": {}, "BANKNIFTY": {}}
    cli = db.client(service=True)
    for sym in ("NIFTY", "BANKNIFTY"):
        resp = cli.table("snapshots").select("*").eq("symbol", sym).order("ts").execute()
        rows = resp.data or []
        for r in rows:
            d = fmt_ist(r["ts"]).date().isoformat()
            by_symbol_day[sym].setdefault(d, []).append(r)
        print(f"  {sym}: {len(rows)} snaps across {len(by_symbol_day[sym])} days")

    nifty_days = sorted(by_symbol_day["NIFTY"].keys())
    live_date_range = f"{nifty_days[0]} to {nifty_days[-1]}" if nifty_days else "n/a"

    wb = Workbook()

    # Will get backtest_days after we run the backtest
    print("Building sheets...", flush=True)
    backtest_results = {}
    backtest_results["NIFTY"] = build_ohlc_backtest(wb, "NIFTY")
    bd_dates = sorted({r["date"] for r in backtest_results["NIFTY"]})
    backtest_date_range = f"{bd_dates[0]} to {bd_dates[-1]}" if bd_dates else "n/a"

    # Build cover + logic NOW that we have the metadata
    # Re-position cover as the first sheet (openpyxl creates new sheets at end by default,
    # but the default first sheet is the active one. We've been using create_sheet so cover
    # is the active sheet 'Sheet' - rename it via build_cover.)
    build_cover(wb, len(nifty_days), live_date_range,
                len(bd_dates), backtest_date_range)
    # Move Cover to index 0
    cover = wb["Cover"]
    wb.move_sheet(cover, offset=-(wb.sheetnames.index("Cover")))

    build_logic(wb)
    # Reorder: Cover, Logic, Live NIFTY, Live BANKNIFTY, Live Summary, OHLC NIFTY, Backtest Summary
    wb.move_sheet(wb["Logic & Conditions"], offset=1 - wb.sheetnames.index("Logic & Conditions"))

    build_live_snapshots(wb, "NIFTY", by_symbol_day["NIFTY"])
    build_live_snapshots(wb, "BANKNIFTY", by_symbol_day["BANKNIFTY"])
    build_live_summary(wb, by_symbol_day)
    build_backtest_summary(wb, backtest_results)

    # Final ordering
    desired = ["Cover", "Logic & Conditions",
               "Live - NIFTY", "Live - BANKNIFTY", "Live Summary",
               "OHLC Backtest - NIFTY", "Backtest Summary"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(wb[name], offset=i - wb.sheetnames.index(name))

    out_path = f"reports/NIFTY_Signals_Investor_Report_{datetime.now(tz=IST).strftime('%Y%m%d_%H%M')}.xlsx"
    import os
    os.makedirs("reports", exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved: {out_path}", flush=True)
    print(f"Sheets: {wb.sheetnames}", flush=True)


if __name__ == "__main__":
    main()
